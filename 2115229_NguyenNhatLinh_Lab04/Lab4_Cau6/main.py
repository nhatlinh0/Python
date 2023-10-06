from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from openpyxl import *
import re

wb = load_workbook('C:\\Users\\Linh\\Desktop\\excel.xlsx')
sheet = wb.active

def excel():
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    sheet.cell(row=1, column=1).value = "MSSV"
    sheet.cell(row=1, column=2).value = "Họ tên"
    sheet.cell(row=1, column=3).value = "Ngày sinh"
    sheet.cell(row=1, column=4).value = "Email"
    sheet.cell(row=1, column=5).value = "SDT"
    sheet.cell(row=1, column=6).value = "Học kỳ"
    sheet.cell(row=1, column=7).value = "Năm học"

def clear():
	
	# clear the content of text entry box
	mssv_field.delete(0, END)
	hoten_field.delete(0, END)
	ngaysinh_field.delete(0, END)
	email_field.delete(0, END)
	sdt_field.delete(0, END)
	hocky_field.delete(0, END)
	namhoc_field.delete(0, END)

def insert():
    current_row = sheet.max_row
    current_column = sheet.max_column
    sheet.cell(row=current_row + 1, column=1).value = mssv_field.get()
    sheet.cell(row=current_row + 1, column=2).value = hoten_field.get()
    sheet.cell(row=current_row + 1, column=3).value = ngaysinh_field.get()
    sheet.cell(row=current_row + 1, column=4).value = email_field.get()
    sheet.cell(row=current_row + 1, column=5).value = sdt_field.get()
    sheet.cell(row=current_row + 1, column=6).value = hocky_field.get()
    sheet.cell(row=current_row + 1, column=7).value = namhoc_field.get()
    wb.save('C:\\Users\\Linh\\Desktop\\excel.xlsx')
    mssv_field.focus_set()
    clear()

root = Tk()
root.configure(background="light green")
root.title("Đăng ký học phần")
root.geometry("500x300")

heading = Label(root, text="THÔNG TIN ĐĂNG KÝ HỌC PHẦN", bg="light green")
mssv = Label(root, text="Mã số sinh viên", bg="light green")
hoten = Label(root, text="Họ tên", bg="light green")
ngaysinh = Label(root, text="Ngày sinh", bg="light green")
email = Label(root, text="Email", bg="light green")
sdt = Label(root, text="Số điện thoại", bg="light green")
hocky = Label(root, text="Học kỳ", bg="light green")
namhoc = Label(root, text="Năm học", bg="light green")
chonmonhoc = Label(root, text="Chọn môn học", bg="light green")

heading.grid(sticky=W, row=0, column=1, padx="80")
mssv.grid(sticky=W, row=1, column=0)
hoten.grid(sticky=W, row=2, column=0)
ngaysinh.grid(sticky=W, row=3, column=0)
email.grid(sticky=W, row=4, column=0)
sdt.grid(sticky=W, row=5, column=0)
hocky.grid(sticky=W, row=6, column=0)
namhoc.grid(sticky=W, row=7, column=0)
chonmonhoc.grid(sticky=W, row=8, column=0)

mssv_field = Entry(root)
hoten_field = Entry(root)
ngaysinh_field = Entry(root)
email_field = Entry(root)
sdt_field = Entry(root)
hocky_field = Entry(root)
n = StringVar()
namhoc_field= ttk.Combobox(root, textvariable=n)
namhoc_field['values'] = ('2022-2023','2023-2024','2024-2025')

# submit
def onClick():
    if(checkMSSV() and checkEmail(email_field.get()) and phoneNumber(sdt_field.get()) and checkHocKy(hocky_field.get()) and checkNgaySinh(ngaysinh_field.get())):
        insert()
    else:
        messagebox.showwarning("Có lỗi", "Vui lòng nhập đúng thông tin")
        

#event khac
regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,7}\b' 
date_pattern = "^[0-9]{1,2}\\/[0-9]{1,2}\\/[0-9]{4}$"

def checkMSSV():
    if(len(mssv_field.get()) < 7):
        return False
    return True

def checkEmail(email):
    if(re.fullmatch(regex, email)):
        return True
    return False

def phoneNumber(num):
    if(num.isdigit() and len(str(num)) == 10):
        return True
    return False

def checkHocKy(num):
    match(num):
        case "1":
            return True
        case "2":
            return True
        case "3":
            return True
        case _n:
            return False

def checkNgaySinh(dat):
    if(re.match(date_pattern, dat)):
        return True
    return False

# 
ltp = Checkbutton(root,background="light green", text="Lập trình Python")
ltj = Checkbutton(root,background="light green", text="Lập trình Java")
cnpm = Checkbutton(root,background="light green", text="Công nghệ phần mềm")
ptudw = Checkbutton(root,background="light green", text="Phát triển ứng dụng web")
btn_dangky = Button(root, background="green", text="Đăng ký", command=onClick)
btn_thoat = Button(root, background="green", text="Thoát", command=quit)

mssv_field.grid(row=1, column=1, ipadx="100")
hoten_field.grid(row=2, column=1, ipadx="100")
ngaysinh_field.grid(row=3, column=1, ipadx="100")
email_field.grid(row=4, column=1, ipadx="100")
sdt_field.grid(row=5, column=1, ipadx="100")
hocky_field.grid(row=6, column=1, ipadx="100")
namhoc_field.grid(row=7, column=1, ipadx="92")
ltp.grid(sticky=W, row=8, column =1, padx=(10,0))
ltj.grid(sticky=E, row=8, column =1, padx=(0,60))
cnpm.grid(sticky=W, row=9, column=1, padx=(10,0))   
ptudw.grid(sticky=E, row=9, column=1)
btn_dangky.grid(row=10, column=1, sticky=W, padx=(60,0))
btn_thoat.grid(row=10, column=1, sticky=E, padx=(0,60))
excel()
root.mainloop()